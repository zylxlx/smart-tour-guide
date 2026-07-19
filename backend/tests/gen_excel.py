"""Generate an Excel report with per-question AI answers, per-category accuracy summary, formatted nicely."""
import sys, os, json, time, urllib.request, urllib.error, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

API = "http://localhost:8001"

CATEGORIES = {
    "景点位置": [
        ("灵山胜境在哪里?", ["无锡","江苏"]),
        ("灵山大佛位于哪里?", ["灵山","无锡","太湖"]),
        ("大佛建在哪个山上?", ["小灵山","灵山","山上"]),
        ("大照壁在哪里?", ["南门","入口","灵山"]),
        ("祥符禅寺位于景区什么位置?", ["灵山","大佛","附近"]),
        ("佛足坛在景区哪个区域?", ["南门","入口","大照壁"]),
        ("百子戏弥勒在哪个广场附近?", ["佛手","广场","弥勒"]),
        ("天下第一掌在哪里?", ["佛手","广场","手掌"]),
        ("灵山精舍位于哪里?", ["灵山","景区","拈花"]),
        ("曼飞龙塔位于哪里?", ["灵山","大佛","附近"]),
        ("五印坛城靠近哪个景点?", ["梵宫","大佛","灵山"]),
        ("拈花湾在哪里?", ["灵山","附近","无锡"]),
        ("五灯湖在哪里?", ["拈花湾","景区","五灯"]),
        ("梵天花海在哪里?", ["灵山","拈花","景区"]),
        ("鹿鸣谷在哪个景区?", ["拈花湾","灵山","鹿"]),
        ("灵山梵宫在景区什么位置?", ["大佛","附近","灵山"]),
        ("菩提大道连接哪些景点?", ["南门","大佛","九龙"]),
        ("胜境广场靠近哪个门?", ["南门","入口","灵山"]),
        ("灵山胜境出口在哪里?", ["祥符禅寺","三圣殿","出口"]),
        ("杏坛广场在什么位置?", ["大佛","附近","灵山"]),
    ],
    "历史文化": [
        ("灵山大佛的建造年代?", ["1997","建成","落成"]),
        ("祥符禅寺有多少年历史?", ["千年","唐代","1300"]),
        ("祥符禅寺最初建于什么时期?", ["唐代","唐朝"]),
        ("祥符禅寺与哪位法师有渊源?", ["玄奘","法师","唐代"]),
        ("灵山什么时候修建的?", ["1994","开工","建设"]),
        ("灵山胜境由哪位大师设计?", ["吴","设计","大师"]),
        ("赵朴初与灵山有什么关系?", ["题写","书法","赵朴初"]),
        ("五智门象征什么?", ["智慧","佛教","五种"]),
        ("三圣殿供奉哪三圣?", ["阿弥陀","观音","大势至"]),
        ("杏坛广场因什么得名?", ["孔子","讲学","杏坛"]),
        ("登云道多少级代表什么含义?", ["108","烦恼","愿望"]),
        ("九龙灌浴讲述什么故事?", ["释迦牟尼","诞生","九龙"]),
        ("九龙灌浴有什么传说?", ["九龙","吐水","神龙"]),
        ("佛足坛有什么圣迹?", ["佛足","足印","佛陀"]),
        ("九龙灌浴的九龙是什么含义?", ["印度","中国","神话","龙"]),
        ("灵山胜境风水中有什么讲究?", ["三山环抱","面朝太湖","风水"]),
        ("灵山精舍是什么?", ["禅意","酒店","住宿"]),
        ("百子戏弥勒的寓意是什么?", ["皆大欢喜","欢喜","快乐"]),
        ("吃素斋有什么意义?", ["禅食","素斋","文化","体验"]),
        ("灵山大照壁上题了什么字?", ["灵山胜境","题字","赵朴初"]),
    ],
    "建筑特色": [
        ("灵山大佛有多高?", ["88米","88","八十八"]),
        ("灵山大佛是什么材质的?", ["青铜","铜"]),
        ("灵山大佛是世界第几大青铜佛像?", ["最大","最高","第一"]),
        ("大照壁全长多少米?", ["39.8","39","四十"]),
        ("大照壁最高处多少米?", ["7米","8.4米","8米"]),
        ("祥符禅寺的大钟有多重?", ["12.8吨","江南第一钟"]),
        ("祥符禅寺的钟叫什么名字?", ["江南第一钟","大钟","铜钟"]),
        ("灵山梵宫面积多大?", ["7.2万","七万"]),
        ("灵山梵宫的穹顶是什么风格?", ["星空","穹顶"]),
        ("灵山梵宫的建筑设计灵感来自哪里?", ["华严经","佛教","经典"]),
        ("灵山梵宫用的是什么琉璃?", ["琉璃","壁画","七彩"]),
        ("灵山梵宫木雕是什么工艺?", ["东阳木雕","木雕","工艺"]),
        ("梵宫有哪些材质?", ["琉璃","木雕","壁画","材质"]),
        ("五印坛城是什么建筑风格?", ["藏传","藏式","藏族","西藏"]),
        ("五印坛城有什么颜色?", ["金","红","金色","红色"]),
        ("五印坛城的金顶象征什么?", ["庄严","神圣","藏传"]),
        ("曼飞龙塔是什么风格?", ["傣族","傣式","少数民族"]),
        ("百子戏弥勒是什么?", ["青铜","雕塑","弥勒"]),
        ("请问灵山大照壁有什么特色?", ["浮雕","佛教","文化","照壁"]),
        ("大佛用什么材料铸造?", ["锡青铜","青铜","铜"]),
    ],
    "游览信息": [
        ("灵山胜境的门票多少钱?", ["210","105","门票"]),
        ("老人门票多少钱?", ["105","半价","优惠"]),
        ("儿童门票多少钱?", ["免票","免费","1.4米"]),
        ("景区开门时间是几点?", ["7点半","7:30","上午","开放"]),
        ("吉祥颂演出时间有哪些?", ["10:35","11:30","14:00","16:00"]),
        ("灵山梵宫演出的名称是什么?", ["吉祥颂","演出"]),
        ("梵宫素斋多少钱一位?", ["50元","五十","素斋"]),
        ("从无锡市区怎么去灵山?", ["88路","89路","公交"]),
        ("从无锡市区打车去灵山多少钱?", ["60","80","打车"]),
        ("登灵山大佛有多少级台阶?", ["216","两百一十六","台阶"]),
        ("大佛脚下可以做什么活动?", ["抱佛脚","祈福","登顶"]),
        ("五印坛城可以做什么?", ["转经筒","祈福","转动"]),
        ("九龙灌浴的圣水可以做什么?", ["接","祈福","圣水"]),
        ("灵山精舍有什么体验项目?", ["抄经","品茶","冥想","早课"]),
        ("景区里给儿童准备了什么活动?", ["抱佛脚","亲子","活动","儿童"]),
        ("灵山精舍可以住宿吗?", ["可以","能","住宿"]),
        ("景区游客最多是什么季节?", ["春秋","春季","秋季"]),
        ("灵山几月份看花最好?", ["春","3月","4月","5月"]),
        ("灵山大佛面向哪个方向?", ["太湖","南","方向"]),
        ("灵山精舍素斋特点?", ["禅食一味","素斋","禅意"]),
    ],
    "路线相关": [
        ("灵山胜境有几条推荐路线?", ["三条","五条","精选","路线"]),
        ("佛教朝圣路线需要多长时间?", ["5小时","五小时","5"]),
        ("自然风光路线需要多长时间?", ["5小时","五小时","5"]),
        ("亲子家庭路线需要多长时间?", ["4小时","四小时","4"]),
        ("历史文化深度路线需要多长时间?", ["6小时","六小时","6"]),
        ("禅修体验路线需要多长时间?", ["4到5","四到五","五","5"]),
        ("佛教朝圣路线包含哪些景点?", ["南门","大照壁","佛足坛","祥符禅寺","大佛","梵宫"]),
        ("灵山有哪些路线主题?", ["佛教","自然","亲子","禅修","历史"]),
        ("亲子路线有哪些亮点?", ["九龙灌浴","百子","梵宫","五印坛城"]),
        ("自然路线有哪些亮点?", ["菩提","太湖","曼飞龙","精舍"]),
        ("历史文化路线有哪些亮点?", ["大照壁","祥符","大佛","梵宫","五印坛城"]),
        ("菩提大道两旁种的是什么树?", ["银杏","树"]),
        ("你能推荐一条适合老人的路线吗?", ["禅修","自然","路线","适合"]),
        ("拈花湾晚上有什么活动?", ["灯光秀","夜市","水幕","禅意"]),
        ("九龙灌浴每天演出几场?", ["每天","多场","场次"]),
        ("菩提大道能看到什么自然风光?", ["太湖","白虎山","龙珠山","自然"]),
        ("拈花湾有什么特点?", ["禅意小镇","客栈","主题","拈花"]),
        ("灵山大佛脚下能看到什么风景?", ["太湖","全景","风光"]),
        ("亲子路线适合什么人?", ["孩子","家庭","亲子","小朋友"]),
        ("灵山精舍适合什么人群?", ["禅修","静心","体验"]),
    ],
}

def ask(q):
    data = json.dumps({"message": q, "session_id": "excel_test"}).encode("utf-8")
    req = urllib.request.Request(f"{API}/api/chat/send", data=data, headers={"Content-Type": "application/json"})
    try:
        r = urllib.request.urlopen(req, timeout=60)
        return json.loads(r.read().decode("utf-8")).get("reply", "")
    except:
        return "[ERROR]"

def check(reply, keywords):
    for w in keywords:
        if w in reply:
            return True
    return False

def build():
    print("Running accuracy test (100 questions)...")
    results = []
    cat_summary = {}

    for cat, qs in CATEGORIES.items():
        correct = 0
        for q, keys in qs:
            reply = ask(q)
            ok = check(reply, keys)
            results.append((cat, q, ", ".join(keys), reply[:120], "✅" if ok else "❌"))
            if ok: correct += 1
            print(f"  {cat[:6]} {correct}/{len(qs)} {q[:30]}...")
            time.sleep(0.3)
        cat_summary[cat] = (correct, len(qs), correct / len(qs) * 100)

    total_correct = sum(v[0] for v in cat_summary.values())
    total_q = sum(v[1] for v in cat_summary.values())
    overall = total_correct / total_q * 100

    # Build Excel
    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "分类别汇总"
    hf = Font(name="黑体", size=14, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    bf = Font(name="宋体", size=12)
    cf = Font(name="黑体", size=12, bold=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    ws.merge_cells("A1:D1")
    ws["A1"] = f"灵山AI禅意导游 — 100题分类别准确率 (总准确率: {overall:.1f}%)"
    ws["A1"].font = Font(name="黑体", size=16, bold=True, color="8B4513")
    ws["A1"].alignment = Alignment(horizontal="center")

    for ci, h in enumerate(["类别", "正确数", "总数", "准确率"], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center"); c.border = thin

    for i, (cat, (corr, tot, acc)) in enumerate(cat_summary.items()):
        row = 4 + i
        for ci, val in enumerate([cat, corr, tot, f"{acc:.1f}%"], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = bf; c.border = thin; c.alignment = Alignment(horizontal="center")

    r = 4 + len(cat_summary)
    for ci, val in enumerate(["总计", total_correct, total_q, f"{overall:.1f}%"], 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = cf; c.border = thin; c.alignment = Alignment(horizontal="center")

    # Bold per-category rows
    for i in range(4, r):
        ws.cell(row=i, column=1).font = Font(name="黑体", size=12, bold=True)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    # ---- Detail sheet ----
    ws2 = wb.create_sheet("详细结果")
    ws2.merge_cells("A1:E1")
    ws2["A1"] = "灵山AI禅意导游 — 100题详细测试结果"
    ws2["A1"].font = Font(name="黑体", size=14, bold=True, color="8B4513")
    ws2["A1"].alignment = Alignment(horizontal="center")

    for ci, h in enumerate(["类别", "问题", "预期答案", "AI回答 (前120字)", "结果"], 1):
        c = ws2.cell(row=3, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.border = thin; c.alignment = Alignment(horizontal="center")

    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    red_fill   = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for i, (cat, q, exp, reply, icon) in enumerate(results):
        row = 4 + i
        for ci, val in enumerate([cat, q, exp, reply, icon], 1):
            c = ws2.cell(row=row, column=ci, value=val)
            c.font = Font(name="宋体", size=10)
            c.border = thin
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if icon == "✅":
                c.fill = green_fill
            else:
                c.fill = red_fill

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 45
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 8

    path = os.path.join(os.path.dirname(__file__), "测试集_分类别准确率.xlsx")
    wb.save(path)
    print(f"\nSaved: {path}")
    print(f"Total accuracy: {total_correct}/{total_q} = {overall:.1f}%")

if __name__ == "__main__":
    build()
