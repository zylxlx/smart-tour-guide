"""景区消费数据分析服务"""
import json
import os
from collections import defaultdict

DATA_PATH = r"D:\文档\20260323113204906\示范景区公开资料包\景点景区旅游数据行为分析数据.xlsx"
CACHE_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "consumption_cache.json")


def _load_and_filter():
    """加载 Excel 并筛选灵山相关数据，缓存为 JSON；无文件时用示例数据"""
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            cache = json.load(f)
        return cache["rows"], cache["total"]

    if not os.path.exists(DATA_PATH):
        return _generate_sample_data()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA_PATH, read_only=True)
    ws = wb[wb.sheetnames[0]]
    keywords = ["灵山", "拈花"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = str(row[4]) if row[4] else ""
        if any(k in name for k in keywords):
            rows.append({
                "id": str(row[0]),
                "name": str(row[1]) if row[1] else "",
                "age": int(row[2]) if row[2] else 0,
                "gender": str(row[3]) if row[3] else "",
                "spot": str(row[4]) if row[4] else "",
                "type": str(row[6]) if row[6] else "",
                "date": str(row[7])[:10] if row[7] else "",
                "stay": float(row[8]) if row[8] else 0,
                "ticket": float(row[9]) if row[9] else 0,
                "food": float(row[10]) if row[10] else 0,
                "shopping": float(row[11]) if row[11] else 0,
                "transport": float(row[12]) if row[12] else 0,
                "entertainment": float(row[13]) if row[13] else 0,
                "total": float(row[14]) if row[14] else 0,
            })

    os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump({"rows": rows, "total": len(rows)}, f, ensure_ascii=False)
    return rows, len(rows)


def _generate_sample_data():
    """生成灵山景区示例消费数据"""
    import random
    random.seed(42)
    spots = ["灵山大佛", "灵山梵宫", "五印坛城", "祥符禅寺", "九龙灌浴", "佛足坛", "大照壁", "菩提大道", "天下第一掌", "百子戏弥勒"]
    genders = ["男", "女"]
    types = ["门票", "餐饮", "购物", "交通", "娱乐"]
    rows = []
    for _ in range(200):
        spot = random.choice(spots)
        rows.append({
            "id": str(random.randint(10000, 99999)),
            "name": f"游客{random.randint(1000,9999)}",
            "age": random.randint(18, 70),
            "gender": random.choice(genders),
            "spot": spot,
            "type": random.choice(types),
            "date": f"2026-0{random.randint(1,6):01d}-{random.randint(1,28):02d}",
            "stay": round(random.uniform(2, 8), 1),
            "ticket": 210.0,
            "food": round(random.uniform(30, 200), 2),
            "shopping": round(random.uniform(0, 500), 2),
            "transport": round(random.uniform(0, 80), 2),
            "entertainment": round(random.uniform(0, 150), 2),
            "total": round(210 + random.uniform(50, 500), 2),
        })
    os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump({"rows": rows, "total": len(rows)}, f, ensure_ascii=False)
    return rows, len(rows)


def get_analysis():
    rows, total = _load_and_filter()

    # 1. 总览
    total_revenue = sum(r["total"] for r in rows)
    total_ticket = sum(r["ticket"] for r in rows)
    total_food = sum(r["food"] for r in rows)
    total_shopping = sum(r["shopping"] for r in rows)
    total_transport = sum(r["transport"] for r in rows)
    total_entertainment = sum(r["entertainment"] for r in rows)
    avg_per_person = total_revenue / total if total > 0 else 0

    # 2. 各景点统计
    spot_stats = defaultdict(lambda: {"count": 0, "revenue": 0.0})
    for r in rows:
        spot_stats[r["spot"]]["count"] += 1
        spot_stats[r["spot"]]["revenue"] += r["total"]
    top_spots = sorted(
        [{"spot": k, **v} for k, v in spot_stats.items()],
        key=lambda x: x["revenue"], reverse=True
    )[:10]

    # 3. 月度趋势
    monthly = defaultdict(lambda: {"revenue": 0.0, "count": 0})
    for r in rows:
        if r["date"]:
            m = r["date"][:7]
            monthly[m]["revenue"] += r["total"]
            monthly[m]["count"] += 1
    monthly_trend = sorted(
        [{"month": k, **v} for k, v in monthly.items()],
        key=lambda x: x["month"]
    )

    # 4. 年龄分布
    age_groups = {"18-25": 0, "26-35": 0, "36-45": 0, "46-55": 0, "56+": 0}
    for r in rows:
        a = r["age"]
        if a <= 25: age_groups["18-25"] += 1
        elif a <= 35: age_groups["26-35"] += 1
        elif a <= 45: age_groups["36-45"] += 1
        elif a <= 55: age_groups["46-55"] += 1
        else: age_groups["56+"] += 1

    # 5. 性别分布
    gender_dist = {"男": 0, "女": 0}
    for r in rows:
        g = r["gender"]
        if g in gender_dist:
            gender_dist[g] += 1

    # 6. 消费类型分布
    category_breakdown = [
        {"name": "门票", "value": round(total_ticket, 2)},
        {"name": "餐饮", "value": round(total_food, 2)},
        {"name": "购物", "value": round(total_shopping, 2)},
        {"name": "交通", "value": round(total_transport, 2)},
        {"name": "娱乐", "value": round(total_entertainment, 2)},
    ]

    return {
        "overview": {
            "total_visitors": total,
            "total_revenue": round(total_revenue, 2),
            "avg_per_person": round(avg_per_person, 2),
            "avg_stay_hours": round(sum(r["stay"] for r in rows) / total, 1) if total > 0 else 0,
        },
        "category_breakdown": category_breakdown,
        "top_spots": top_spots,
        "monthly_trend": monthly_trend,
        "age_groups": age_groups,
        "gender_dist": gender_dist,
    }
